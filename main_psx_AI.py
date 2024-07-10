import datetime
import logging
import os
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import sqlalchemy
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

from analysis_functions import analyze_symbol
from convert_excel_to_lists import CUSTUM, KMIALL, KMI100, KMI30, MYLIST, QSE, CRYPTO
from email_functions import send_email
from telegram_message import send_telegram_message

# Set up logging
logging.basicConfig(filename='analysis_log.txt', level=logging.ERROR)

# Global constants
VOLUME_THRESHOLD = 100000
MIN_VOLUME = 5000
AO_THRESHOLD = 0
RSI_THRESHOLD = 30

def get_analysis_type():
    return "D"  # Placeholder function

def get_recommendation_filter():
    return "STRONG_BUY"  # Placeholder function

def get_symbol_selection(country_selection):
    if country_selection == "P":
        return "KMI100"  # Placeholder symbol selection for Pakistan
    else:
        return "QSE"  # Placeholder symbol selection for Qatar

def main(country_selection):
    analysis_type = get_analysis_type()
    recommendation_filter = get_recommendation_filter()
    current_datetime = datetime.datetime.now()
    count = 1

    if country_selection == "P":
        screener_selection, exchange_selection = "PAKISTAN", "PSX"
        base_url = {
            'charts': "https://www.tradingview.com/chart/ZMYE714n/?symbol=PSX%3A",
            'finance': "https://www.tradingview.com/symbols/PSX-",
            'tech': "https://www.tradingview.com/symbols/PSX-"
        }
    else:
        screener_selection, exchange_selection = "QATAR", "QSE"
        base_url = {
            'charts': "https://www.tradingview.com/chart/ZMYE714n/?symbol=QSE%3A",
            'finance': "https://www.tradingview.com/symbols/QSE-",
            'tech': "https://www.tradingview.com/symbols/QSE-"
        }

    symbol_selection = get_symbol_selection(country_selection)
    symbols = {
        "KMI30": KMI30,
        "KMI100": KMI100,
        "MYLIST": MYLIST,
        "CRYPTO": CRYPTO,
        "CUSTUM": CUSTUM,
        "QSE": QSE
    }.get(symbol_selection, KMIALL)

    analyzed_data = []
    strong_buy_symbols = []
    buy_symbols = []
    sell_symbols = []
    AO_symbols = []

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [
            executor.submit(
                analyze_symbol, symbol, analysis_type, screener_selection, exchange_selection,
                base_url['charts'], base_url['finance'], base_url['tech']
            )
            for symbol in symbols
        ]

        for future in futures:
            result = future.result()
            if result:
                (symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume, ADX, RSI, RSI_Last, AO, 
                 Change, avg_support, avg_resistance, chart_url, finance_url, tech_url) = result
                
                summary = summary.get('RECOMMENDATION')
                print(f"{count}: Symbol: {symbol} - {summary}")
                print(f"{analysis_type} Sell_Signal:{Sell_Signal}, Neutral_Signal:{Neutral_Signal}, "
                      f"Buy_Signal:{Buy_Signal}, CLOSE: {Close}, Volume: {Volume}, ADX: {ADX}, RSI: {RSI}, "
                      f"LAST RSI: {RSI_Last}, AO: {AO}, %Change(D): {Change}\n")
                
                count += 1

                data_entry = [
                    current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                    ADX, RSI, RSI_Last, AO, Change, avg_support, avg_resistance, chart_url, finance_url, tech_url
                ]

                analyzed_data.append(data_entry)

                if (summary in ["STRONG_BUY", "BUY", "NEUTRAL"] and isinstance(Volume, (int, float)) and Volume > VOLUME_THRESHOLD and isinstance(AO, (int, float)) and AO > AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI > RSI_THRESHOLD):
                    strong_buy_symbols.append(data_entry)

                if summary == "BUY" and isinstance(Volume, (int, float)) and Volume > MIN_VOLUME and isinstance(AO, (int, float)) and AO > AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI > RSI_THRESHOLD:
                    buy_symbols.append(data_entry)

                if summary == "SELL" and isinstance(AO, (int, float)) and AO < AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI < RSI_THRESHOLD:
                    sell_symbols.append(data_entry)

                if summary in ["NEUTRAL", "SELL", "STRONG_SELL"] and isinstance(AO, (int, float)) and AO < AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI < RSI_THRESHOLD:
                    AO_symbols.append(data_entry)

    save_to_excel(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection, recommendation_filter)
    save_to_database(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection)

    send_notifications(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, analysis_type, symbol_selection, recommendation_filter)

def save_to_excel(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection, recommendation_filter):
    excel_file_path = f"{analysis_type}-Advance_Technical_Analysis_{symbol_selection}_{recommendation_filter}.xlsx"
    
    subset_columns = [
        'Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 
        'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance'
    ]
    subset_columns_drop = subset_columns[1:]

    dataframes = {
        'All Symbols': pd.DataFrame(analyzed_data, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'Recommended_Symbols': pd.DataFrame(strong_buy_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'Buy_Symbols': pd.DataFrame(buy_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'Sell_Symbols': pd.DataFrame(sell_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'AO_Symbols': pd.DataFrame(AO_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals'])
    }

    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        for sheet_name in dataframes:
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 20
            column_J_ranges = worksheet['J']
            column_L_ranges = worksheet['L']

            for cell in column_J_ranges:
                cell.font = Font(color="FF0000")

            for cell in column_L_ranges:
                cell.font = Font(color="FF0000")

            highlight_format_bullish = PatternFill(start_color="75AA74", end_color="75AA74", fill_type="solid")
            highlight_format_bearish = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            highlight_format_neutral = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            highlight_format_exit = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            dxf_bullish = DifferentialStyle(fill=highlight_format_bullish)
            dxf_bearish = DifferentialStyle(fill=highlight_format_bearish)
            dxf_neutral = DifferentialStyle(fill=highlight_format_neutral)
            dxf_exit = DifferentialStyle(fill=highlight_format_exit)

            rule_bullish = Rule(type="expression", dxf=dxf_bullish)
            rule_bullish.formula = ["$J1>50", "$L1>0"]
            rule_bearish = Rule(type="expression", dxf=dxf_bearish)
            rule_bearish.formula = ["$J1<50", "$L1<0"]
            rule_neutral = Rule(type="expression", dxf=dxf_neutral)
            rule_neutral.formula = ["$J1>50", "$L1>0"]
            rule_exit = Rule(type="expression", dxf=dxf_exit)
            rule_exit.formula = ["$E1>7"]

            worksheet.conditional_formatting.add("A1:R1000", rule_bullish)
            worksheet.conditional_formatting.add("E1:E1000", rule_exit)

def save_to_database(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection):
    database_url = f'sqlite:///{symbol_selection}_{analysis_type}_data.db'
    engine = sqlalchemy.create_engine(database_url)

    dataframes = {
        'all_symbols': pd.DataFrame(analyzed_data, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'recommended_symbols': pd.DataFrame(strong_buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'buy_symbols': pd.DataFrame(buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'sell_symbols': pd.DataFrame(sell_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'AO_symbols': pd.DataFrame(AO_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    }

    for table_name, df in dataframes.items():
        df.to_sql(table_name, engine, if_exists='append', index=False)

def send_notifications(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, analysis_type, symbol_selection, recommendation_filter):
    today_date = datetime.datetime.now().date()
    
    df_strong_buy = pd.DataFrame(strong_buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    df_buy = pd.DataFrame(buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    df_sell = pd.DataFrame(sell_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    
    today_Strong_buy = df_strong_buy[df_strong_buy['Date and Time'].dt.date == today_date]
    today_buy = df_buy[df_buy['Date and Time'].dt.date == today_date]
    today_sell = df_sell[df_sell['Date and Time'].dt.date == today_date]
    
    custom_message_strong_buy = f"Strong Buy Conditions: Volume >{VOLUME_THRESHOLD},RSI>{RSI_THRESHOLD}, AO >{AO_THRESHOLD} and {recommendation_filter} from Trading view."
    custom_message_buy = f"Buy Conditions: AO > {AO_THRESHOLD} and Buy from Trading view and RSI>{RSI_THRESHOLD}"
    custom_message_sell = f"Sell Conditions: AO < {AO_THRESHOLD} and Sell from Trading view and RSI<{RSI_THRESHOLD}"
    
    today_Strong_buy_mail = today_Strong_buy.to_string(index=True, justify='left', col_space=10)
    today_buy_mail = today_buy.to_string(index=True, justify='left', col_space=10)
    today_sell_mail = today_sell.to_string(index=True, justify='left', col_space=10)
    
    body = f"\n\n\n{custom_message_sell}\n{today_sell_mail}\n\n\n{custom_message_buy}\n{today_buy_mail}\n\n\n{custom_message_strong_buy}\n{today_Strong_buy_mail}"
    
    attachment_path = f"{analysis_type}-Advance_Technical_Analysis_{symbol_selection}_{recommendation_filter}.xlsx"
    subject = f"{analysis_type}-{symbol_selection}- {recommendation_filter}-Technical_Analysis_"
    
    try:
        send_email(subject, body, attachment_path)
        send_telegram_message(body)
    except Exception as e:
        print(f"Error sending email: {str(e)}")

if __name__ == "__main__":
    current_day = datetime.datetime.now().weekday()
    current_month_day = datetime.datetime.now().day

    if current_day == 5:
        analysis_type = "W"
    elif current_month_day == 1:
        analysis_type = "M"
    else:
        analysis_type = "D"

    main("P")
    print("PSX Analysis complete")
    time.sleep(5)

    main("Q")
    print("QSE Analysis complete")
    time.sleep(5)

    print("All Analysis complete")
    time.sleep(5)
