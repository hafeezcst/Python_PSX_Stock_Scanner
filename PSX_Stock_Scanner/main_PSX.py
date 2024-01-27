import datetime
import logging
import os
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import sqlalchemy
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

from analysis_functions import analyze_symbol
from convert_excel_to_lists import CUSTUM, KMIALL, KMI100, KMI30, MYLIST, QSE
from email_functions import send_email
from get_analysis_type import get_analysis_type
from get_recommendation_filter import get_recommendation_filter
from get_symbol_selection import get_symbol_selection

# Set up logging
logging.basicConfig(filename='analysis_log.txt', level=logging.ERROR)

# Global constants
# Set the volume threshold to 1,000,000
VOLUME_THRESHOLD = 100000

# Set the minimum volume to 50,00
MIN_VOLUME = 5000

# Set the AO threshold to 0
AO_THRESHOLD = 0


def main():
    # Ask the user to select the analysis type and convert the input to uppercase
    analysis_type = get_analysis_type()
    # If the user didn't provide any input
    # set default analysis type to "D"
    if analysis_type != "D":
        # Print the selected analysis type
        print(f"Selected analysis type: {analysis_type}")
    # Ask the user to select the analysis type and convert the input to uppercase
    country_selection = input(
        "Select analysis type (P for Pakistan, Q for Qatar, press Enter for default PAKISTAN):").upper()
    # If the user didn't provide any input
    # set default country to PAKISTAN
    country_selection = "P" if country_selection == "" else country_selection
    if country_selection == "P":
        # Wait for 2 seconds
        time.sleep(2)
        # Set the analysis type to "D" by default
        screener_selection, exchange_selection = "PAKISTAN", "PSX"
        BASE_URL_CHARTS = f"https://www.tradingview.com/chart/ZMYE714n/?symbol=PSX%3A"
        BASE_URL_FINANCE = f"https://www.tradingview.com/symbols/PSX-"
        BASE_URL_TECH = f"https://www.tradingview.com/symbols/PSX-"
        # Print the selected analysis type
        print(f"Selected country : {screener_selection}")
        print(f"Selected exchange: {exchange_selection}")
    elif country_selection == "Q":
        # Wait for 2 seconds
        time.sleep(2)
        # Set the analysis type to "D" by default
        screener_selection, exchange_selection = "QATAR", "QSE"
        BASE_URL_CHARTS = f"https://www.tradingview.com/chart/ZMYE714n/?symbol=QSE%3A"
        BASE_URL_FINANCE = f"https://www.tradingview.com/symbols/QSE-"
        BASE_URL_TECH = f"https://www.tradingview.com/symbols/QSE-"
        # Print the selected analysis type
        print(f"Selected country : {screener_selection}")
        print(f"Selected exchange: {exchange_selection}")

    # Ask the user to select a recommendation filter and convert the input to uppercase
    recommendation_filter = get_recommendation_filter()

    # Set default analysis type to "M" if the user presses Enter without entering a choice
    current_datetime = datetime.datetime.now()  # Include time
    count = 1

    # Create a thread pool
    # Adjust max_workers as needed
    with ThreadPoolExecutor(max_workers=6) as executor:
        # Submit tasks to the thread pool
        # Ask the user to select the symbol list and convert the input to uppercase
        symbol_selection = get_symbol_selection()
        if symbol_selection != "KMIALL":
            print(f"Selected symbol List: {symbol_selection}")
        # If symbol_selection is "KMI30"
        if symbol_selection == "KMI30":
            # Set psx_symbols to the value of KMI30
            psx_symbols = KMI30
        # Else, if symbol_selection is "KMI100"
        elif symbol_selection == "KMI100":
            # Set psx_symbols to the value of KMI100
            psx_symbols = KMI100
        # Else, if symbol_selection is "MYLIST"
        elif symbol_selection == "MYLIST":
            # Set psx_symbols to the value of MYLIST
            psx_symbols = MYLIST
        # Else, if symbol_selection is "CUSTUM"
        elif symbol_selection == "CUSTUM":
            # Set psx_symbols to the value of CUSTUM
            psx_symbols = CUSTUM
        elif symbol_selection == "QSE":
            # Set psx_symbols to the value of CUSTUM
            psx_symbols = QSE
        # Else, if symbol_selection is anything else
        else:
            # Set psx_symbols to the value of KMIALL
            psx_symbols = KMIALL
        # Define the base URLs

        # For each symbol in psx_symbols
        futures = [
            # Submit a task to the executor to analyze the symbol
            executor.submit(
                analyze_symbol,  # The function to execute
                symbol,  # The symbol to analyze
                analysis_type,  # The type of analysis to perform
                screener_selection,  # The screener to use
                exchange_selection,  # The exchange to use
                BASE_URL_CHARTS,  # The base URL for chart data
                BASE_URL_FINANCE,  # The base URL for financial data
                BASE_URL_TECH  # The base URL for technical data
            )
            for symbol in psx_symbols  # The list of symbols to analyze
        ]

        # Initialize an empty list to store all analyzed data
        analyzed_data = []
        # Initialize an empty list to store symbols with a "STRONG_BUY" recommendation
        strong_buy_symbols = []
        # Initialize an empty list to store symbols with a "BUY" recommendation
        buy_symbols = []
        # Initialize an empty list to store symbols with a "SELL" recommendation
        sell_symbols = []
        # Initialize an empty list to store symbols with a "AO" recommendation
        AO_symbols = []
        # For each future in the list of futures

        for future in futures:
            # Get the result of the future
            result = future.result()

            # If the result is not None
            if result:
                # Unpack the result into multiple variables
                symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume, ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS, BASE_URL_FINANCE, BASE_URL_TECH = result

                # Print the count and symbol
                print(f"{count}:  Symbol: {symbol}")

                # Get the 'RECOMMENDATION' from the summary
                summary = summary.get('RECOMMENDATION')

                # Print the summary
                print(summary)

                # Print various details about the symbol
                print(
                    f"{analysis_type} Sell_Signal:{Sell_Signal},Neutral_Signal:{Neutral_Signal},Buy_Signal:{Buy_Signal},CLOSE: {Close} Volume: {Volume} ADX:{ADX} RSI: {RSI} LAST RSI: {RSI_Last} AO: {AO} %Change(D): {Change}")

                # Print an empty line
                print()

                # Increment the count
                count += 1

                # If Volume and AO are not None

                # If Volume is greater than min_volume
                if symbol not in ["KMI30", "KMI100", "ALLSHR","GNRI"] and Volume is not None and Volume > MIN_VOLUME:
                    # Append the result to the analyzed_data list
                    analyzed_data.append([
                        current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                        ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS,
                        BASE_URL_FINANCE, BASE_URL_TECH
                    ])
                else:
                    # Append the result to the analyzed_data list
                    analyzed_data.append([
                        current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                        ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS,
                        BASE_URL_FINANCE, BASE_URL_TECH
                    ])
                # Check if the recommendation is "user defined" and the volume is greater than the threshold
                if (
                        summary == recommendation_filter and Volume is not None and Volume > VOLUME_THRESHOLD) and AO > AO_THRESHOLD or (
                        symbol in ["KSE30", "KSE100", "ALLSHR", "GNRI"]):
                    strong_buy_symbols.append([
                        current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                        ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS,
                        BASE_URL_FINANCE, BASE_URL_TECH
                    ])

                # Check if the recommendation is "fixed buy" and the volume is greater than the threshold
                if summary == "BUY" and Volume is not None and Volume > MIN_VOLUME and AO > AO_THRESHOLD:
                    buy_symbols.append([
                        current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                        ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS,
                        BASE_URL_FINANCE, BASE_URL_TECH
                    ])
                # Check if the recommendation is "fixed buy" and the volume is greater than the threshold
                if summary == "SELL" and AO < AO_THRESHOLD:
                    sell_symbols.append([
                        current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                        ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS,
                        BASE_URL_FINANCE, BASE_URL_TECH
                    ])
                # Check if the recommendation is "fixed buy" and the volume is greater than the threshold
                if summary in ["STRONG_BUY", "BUY", "NEUTRAL", "SELL", "STRONG_SELL"] and AO > AO_THRESHOLD:
                    AO_symbols.append([
                        current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                        ADX, RSI, RSI_Last, AO, Change, average_support, average_resistance, BASE_URL_CHARTS,
                        BASE_URL_FINANCE, BASE_URL_TECH
                    ])

                # Set the file name with the analysis date as a postfix
                excel_file_path = f"{analysis_type}-Advance_Technical_Analysis_{symbol_selection}_{recommendation_filter}.xlsx"
        # Define a list of column names
        subset_columns = [
            'Date and Time',  # The date and time of the analysis
            'Symbol',  # The symbol of the stock
            'Summary',  # The summary of the analysis
            # The closing price, with the analysis type as a prefix
            f'{analysis_type} Close',
            'Sell',  # The sell signal
            'Neutral',  # The neutral signal
            'Buy',  # The buy signal
            'Volume',  # The trading volume
            'ADX',  # The Average Directional Index
            'RSI',  # The Relative Strength Index
            'Last RSI',  # The last value of the Relative Strength Index
            'AO',  # The Awesome Oscillator
            '%Change(D)',  # The daily percentage change
            'Support',  # The support level
            'Resistance'  # The resistance level
        ]
        subset_columns_drop = [
            'Symbol',  # The symbol of the stock
            'Summary',  # The summary of the analysis
            # The closing price, with the analysis type as a prefix
            f'{analysis_type} Close',
            'Sell',  # The sell signal
            'Neutral',  # The neutral signal
            'Buy',  # The buy signal
            'Volume',  # The trading volume
            'ADX',  # The Average Directional Index
            'RSI',  # The Relative Strength Index
            'Last RSI',  # The last value of the Relative Strength Index
            'AO',  # The Awesome Oscillator
            '%Change(D)',  # The daily percentage change
            'Support',  # The support level
            'Resistance'  # The resistance level
        ]
        # Check if the file exists
    if os.path.exists(excel_file_path):
        print(f"Reading data from {excel_file_path}")
        # Read existing data from the file
        # Read data from each sheet into separate DataFrames
        while True:
            try:
                if os.path.exists(excel_file_path):
                    with pd.ExcelFile(excel_file_path) as xls:
                        if 'All Symbols' in xls.sheet_names:
                            df_all_symbols = pd.read_excel(
                                excel_file_path, sheet_name='All Symbols')
                        else:
                            df_all_symbols = pd.DataFrame(
                                columns=subset_columns)
                        if 'Recommended_Symbols' in xls.sheet_names:
                            df_strong_buy_symbols = pd.read_excel(
                                excel_file_path, sheet_name='Recommended_Symbols')
                        else:
                            df_strong_buy_symbols = pd.DataFrame(
                                columns=subset_columns)
                        if 'Buy_Symbols' in xls.sheet_names:
                            df_buy_symbols = pd.read_excel(
                                excel_file_path, sheet_name='Buy_Symbols')
                        else:
                            df_buy_symbols = pd.DataFrame(
                                columns=subset_columns)
                        if 'Sell_Symbols' in xls.sheet_names:
                            df_sell_symbols = pd.read_excel(
                                excel_file_path, sheet_name='Sell_Symbols')
                        else:
                            df_sell_symbols = pd.DataFrame(
                                columns=subset_columns)
                        if 'AO_Symbols' in xls.sheet_names:
                            df_AO_symbols = pd.read_excel(
                                excel_file_path, sheet_name='AO_Symbols')
                        else:
                            df_AO_symbols = pd.DataFrame(
                                columns=subset_columns)
                else:
                    df_all_symbols = pd.DataFrame(columns=subset_columns)
                    df_strong_buy_symbols = pd.DataFrame(
                        columns=subset_columns)
                    df_buy_symbols = pd.DataFrame(columns=subset_columns)
                    df_sell_symbols = pd.DataFrame(columns=subset_columns)
                    df_AO_symbols = pd.DataFrame(columns=subset_columns)
                break  # If the file opens successfully, break the loop
            except PermissionError:
                print(f"Waiting for file to close: {excel_file_path}")
                time.sleep(5)  # Wait for 5 seconds before trying again

        # Concatenate the new data with the existing data
        df_all = pd.concat([df_all_symbols, pd.DataFrame(analyzed_data,
                            columns=['Date and Time', 'Symbol', 'Summary',
                                                                f'{analysis_type} Close', 'Sell',
                                                                'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                                                                'Last RSI', 'AO', '%Change(D)', 'Support',
                                                                'Resistance',
                                                                'Charts', 'Financials',
                                                                'Technicals'])]).drop_duplicates(
            subset=subset_columns_drop, keep='last')

        df_strong_buy = pd.concat([df_strong_buy_symbols, pd.DataFrame(strong_buy_symbols,
                            columns=['Date and Time', 'Symbol',
                                                                'Summary',
                                                                f'{analysis_type} Close', 'Sell',
                                                                'Neutral', 'Buy', 'Volume', 'ADX',
                                                                'RSI',
                                                                'Last RSI', 'AO', '%Change(D)',
                                                                'Support', 'Resistance',
                                                                'Charts', 'Financials',
                                                                'Technicals'])]).drop_duplicates(
            subset=subset_columns_drop, keep='last')

        df_buy = pd.concat([df_buy_symbols, pd.DataFrame(buy_symbols,
                            columns=['Date and Time', 'Symbol',
                                                                  'Summary',
                                                                  f'{analysis_type} Close',
                                                                  'Sell', 'Neutral', 'Buy',
                                                                  'Volume', 'ADX', 'RSI',
                                                                  'Last RSI', 'AO',
                                                                  '%Change(D)', 'Support', 'Resistance', 'Charts',
                                                                  'Financials',
                                                                  'Technicals'])]).drop_duplicates(
            subset=subset_columns_drop, keep='last')
        df_sell = pd.concat([df_sell_symbols, pd.DataFrame(sell_symbols,
                            columns=['Date and Time', 'Symbol',
                                                                    'Summary',
                                                                    f'{analysis_type} Close',
                                                                    'Sell', 'Neutral', 'Buy',
                                                                    'Volume', 'ADX', 'RSI',
                                                                    'Last RSI', 'AO',
                                                                    '%Change(D)', 'Support', 'Resistance',
                                                                    'Charts',
                                                                    'Financials',
                                                                    'Technicals'])]).drop_duplicates(
            subset=subset_columns_drop, keep='last')
        df_AO = pd.concat([df_AO_symbols, pd.DataFrame(AO_symbols,
                            columns=['Date and Time', 'Symbol',
                                                                'Summary',
                                                                f'{analysis_type} Close',
                                                                'Sell', 'Neutral', 'Buy',
                                                                'Volume', 'ADX', 'RSI',
                                                                'Last RSI', 'AO',
                                                                '%Change(D)', 'Support', 'Resistance', 'Charts',
                                                                'Financials',
                                                                'Technicals'])]).drop_duplicates(
            subset=subset_columns_drop, keep='last')
    else:
        # Create a new file
        subset_columns = ['Symbol', 'Summary', f'{analysis_type} Close', 'Sell',
                          'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                          'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance']
        df_all = pd.DataFrame(analyzed_data,
                            columns=['Date and Time', 'Symbol', 'Summary',
                                       f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                                       'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance',
                                       'Charts', 'Financials', 'Technicals'])

        df_strong_buy = pd.DataFrame(strong_buy_symbols,
                            columns=['Date and Time', 'Symbol', 'Summary',
                                              f'{analysis_type} Close', 'Sell',
                                              'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                                              'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance',
                                              'Charts', 'Financials', 'Technicals'])
        df_buy = pd.DataFrame(buy_symbols,
                            columns=['Date and Time', 'Symbol', 'Summary',
                                       f'{analysis_type} Close', 'Sell',
                                       'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                                       'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance',
                                       'Charts', 'Financials', 'Technicals'])
        df_sell = pd.DataFrame(sell_symbols,
                            columns=['Date and Time', 'Symbol', 'Summary',
                                        f'{analysis_type} Close', 'Sell',
                                        'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                                        'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance',
                                        'Charts', 'Financials', 'Technicals'])
        df_AO = pd.DataFrame(AO_symbols,
                            columns=['Date and Time', 'Symbol', 'Summary',
                                      f'{analysis_type} Close', 'Sell',
                                      'Neutral', 'Buy', 'Volume', 'ADX', 'RSI',
                                      'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance',
                                      'Charts', 'Financials', 'Technicals'])

    # Define the database connection URL
    # The URL is in the format 'sqlite:///<database_name>.db'
    # The database name is a combination of the symbol_selection and analysis_type variables
    database_url = f'sqlite:///{symbol_selection}_{analysis_type}_data.db'

    # Create a database engine
    # The engine is an instance of a SQLAlchemy Engine, which is the starting point for any SQLAlchemy application
    # It's the home base for the actual database and its DBAPI, delivered to the SQLAlchemy application through a connection pool and a Dialect
    engine = sqlalchemy.create_engine(database_url)

    # Write DataFrames to the database
    # If at least one of the DataFrames is not None
    if df_all is not None or df_strong_buy is not None or df_buy is not None or df_sell is not None:
        # Write df_all to a SQL table named 'all_symbols'
        df_all.to_sql('all_symbols', engine, if_exists='append', index=False)
        # Write df_strong_buy to a SQL table named 'recommended_symbols'
        df_strong_buy.to_sql('recommended_symbols', engine,
                             if_exists='append', index=False)
        # Write df_buy to a SQL table named 'buy_symbols'
        df_buy.to_sql('buy_symbols', engine, if_exists='append', index=False)
        # Write df_sell to a SQL table named 'sell_symbols'
        df_sell.to_sql('sell_symbols', engine, if_exists='append', index=False)
        # Write df_AO to a SQL table named 'AO_symbols'
        df_AO.to_sql('AO_symbols', engine, if_exists='append', index=False)

    else:
        # If all the DataFrames are None, print a message
        print("df_all is None")

    # Save the Excel file
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Write df_all to an Excel sheet named 'All Symbols', without the index
        df_all.to_excel(writer, sheet_name='All Symbols', index=False)
        # Write df_strong_buy to an Excel sheet named 'Recommended_Symbols', without the index
        df_strong_buy.to_excel(
            writer, sheet_name='Recommended_Symbols', index=False)
        # Write df_buy to an Excel sheet named 'Buy_Symbols', without the index
        df_buy.to_excel(writer, sheet_name='Buy_Symbols', index=False)
        # Write df_sell to an Excel sheet named 'Sell_Symbols', without the index
        df_sell.to_excel(writer, sheet_name='Sell_Symbols', index=False)
        # Write df_AO to an Excel sheet named 'AO_Symbols', without the index
        df_AO.to_excel(writer, sheet_name='AO_Symbols', index=False)

        # Get the worksheet named 'All Symbols'
        worksheet_all = writer.sheets['All Symbols']

        # Get the worksheet named 'Recommended_Symbols'
        worksheet_strong_buy = writer.sheets['Recommended_Symbols']

        # Get the worksheet named 'Buy_Symbols'
        worksheet_buy = writer.sheets['Buy_Symbols']

        # Get the worksheet named 'Sell_Symbols'
        worksheet_sell = writer.sheets['Sell_Symbols']
        # Get the worksheet named 'AO_Symbols'
        worksheet_AO = writer.sheets['AO_Symbols']
        # Set the column width and format
        worksheet_all.column_dimensions['A'].width = 20
        worksheet_strong_buy.column_dimensions['A'].width = 20
        worksheet_buy.column_dimensions['A'].width = 20
        worksheet_sell.column_dimensions['A'].width = 20
        worksheet_AO.column_dimensions['A'].width = 20
        # Set the column width and format
        # Get the column R range
        column_J_ranges = [worksheet_all['J'], worksheet_strong_buy['J'], worksheet_buy['J'],
                           worksheet_sell['J'],
                           worksheet_AO['J']]
        column_L_ranges = [worksheet_all['L'], worksheet_strong_buy['L'], worksheet_buy['L'],
                           worksheet_sell['L'],
                           worksheet_AO['L']]

        # Set the font color to red for all J columns
        for column_J_range in column_J_ranges:
            for cell in column_J_range:
                cell.font = Font(color="FF0000")

        # Set the font color to red for all L columns
        for column_L_range in column_L_ranges:
            for cell in column_L_range:
                cell.font = Font(color="FF0000")

        # Create a fill pattern
        highlight_format_bullish = PatternFill(
            start_color="75AA74", end_color="75AA74", fill_type="solid")
        # Create a fill pattern
        highlight_format_bearish = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        highlight_format_neutral = PatternFill(
            start_color="98FB98", end_color="98FB98", fill_type="solid")
        highlight_format_exit = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        # Create a differential style
        dxf_bullish = DifferentialStyle(fill=highlight_format_bullish)
        dxf_bearish = DifferentialStyle(fill=highlight_format_bearish)
        dxf_neutral = DifferentialStyle(fill=highlight_format_neutral)
        dxf_exit = DifferentialStyle(fill=highlight_format_exit)
        # Create a rule bullish
        rule_bullish = Rule(type="expression", dxf=dxf_bullish)
        rule_bullish.formula = ["$J1>50"] and ["$L1>0"]
        # Create a rule bearish
        rule_bearish = Rule(type="expression", dxf=dxf_bearish)
        rule_bearish.formula = ["$J1<50"] and ["$L1<0"]
        # Create a rule neutral
        rule_neutral = Rule(type="expression", dxf=dxf_neutral)
        rule_neutral.formula = ["$J1>50"] and ["$L1>0"]
        # Create a rule Exit
        rule_exit = Rule(type="expression", dxf=dxf_exit)
        rule_exit.formula = ["$E1>7"]

        # Add the rule to the worksheet
        worksheet_all.conditional_formatting.add("A1:R1000", rule_bullish)
        worksheet_all.conditional_formatting.add("E1:E1000", rule_exit)
        worksheet_strong_buy.conditional_formatting.add(
            "A1:R1000", rule_bullish)
        worksheet_buy.conditional_formatting.add("A1:R1000", rule_bullish)
        worksheet_sell.conditional_formatting.add("A1:R1000", rule_bearish)
        worksheet_AO.conditional_formatting.add("A1:R1000", rule_neutral)

        # Save the Excel file
        print(f"Data exported to {excel_file_path}")

        # Wait until the file size stabilizes (e.g., for 5 seconds)
        initial_size = os.path.getsize(excel_file_path)
        time.sleep(5)  # Adjust the delay as needed
    while True:
        current_size = os.path.getsize(excel_file_path)
        if current_size == initial_size:
            break  # File size has stabilized
        initial_size = current_size
        time.sleep(2)  # Wait and check again
        # Explicitly close the Excel write
        # Add a short delay (e.g., 1 second) before sending the email
        time.sleep(3)

        # send email
        # Assuming the correct DataFrame is named df_strong_buy:
        # Sort df_strong_buy by 'Date and Time' in descending order
        recommended_symbols = df_strong_buy.sort_values(
            by='Date and Time', ascending=False)
        # Sort df_buy by 'Date and Time' in descending order
        buy_symbols = df_buy.sort_values(by='Date and Time', ascending=False)
        # Filter the symbols for the current date
        today_date = datetime.datetime.now().date()
        one_hours_ago = datetime.datetime.now() - datetime.timedelta(hours=1)
        today_Strong_buy = recommended_symbols[recommended_symbols[
            'Date and Time'].dt.date == today_date] if not recommended_symbols.empty else pd.DataFrame()
        # Apply custom column filter

        # Calculate the suggested number of shares to purchase
        investment_amount = 4000000
        num_shares = len(today_Strong_buy[today_Strong_buy['Volume'] > 0])
        if num_shares > 0:
            investment_amount_per_share = round(investment_amount / num_shares)
            suggested_shares = today_Strong_buy.apply(
                lambda row: round(investment_amount_per_share / row[f'{analysis_type} Close']) if row[
                    'Volume'] > 0 else 0,
                axis=1)
        else:
            suggested_shares = pd.Series()

        today_Strong_buy = today_Strong_buy.copy()
        today_Strong_buy.loc[:, 'Suggested Shares'] = suggested_shares

        # Convert the filtered data to a string

        custom_message = f"Strong Buy Conditions: Volume >{VOLUME_THRESHOLD},RSI>50, AO >{AO_THRESHOLD} and{recommendation_filter} from Trading view.Suggested Shares are calculated based on {investment_amount} investment amount. min volume >{MIN_VOLUME} "
        body = f"{custom_message}\n\n{today_Strong_buy.to_string ( index=True, justify='left', col_space=10, header=True, na_rep='NaN', formatters=None )}"
        # Define the subject of the email
        attachment_path = excel_file_path
        subject = f"{analysis_type}-{symbol_selection}- {recommendation_filter}-Technical_Analysis_"
        send_email(subject, body, attachment_path)
        # Sort the Excel file by symbol and date and time
        # Open the file using the default program associated with Excel files
        subprocess.Popen([excel_file_path], shell=True)
        # Wait for 5 seconds before running the analysis again
        # today_buy= buy_symbols[buy_symbols['Date and Time'].dt.date == today_date] if not buy_symbols.empty else pd.DataFrame()


# If this script is the main module (i.e., it's not being imported by another script)
if __name__ == "__main__":
    # Initialize an empty list to store all analyzed data
    analyzed_data = []
    # Initialize an empty list to store symbols with a "STRONG_BUY" recommendation
    strong_buy_symbols = []
    # Initialize an empty list to store symbols with a "BUY" recommendation
    buy_symbols = []
    # Call the main function to run the main analysis
    main()
