try:
    import pandas as pd
except ImportError:
    print("Please install the 'pandas' library.")
    raise
import logging
from openpyxl.styles import PatternFill
from datetime import date

# Setup basic configuration for logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def read_and_analyze_stock_data(file_path):
    try:
        # Read the Excel file from the specified sheet
        df = pd.read_excel(file_path, sheet_name='All Symbols')
        logging.info("Excel file read successfully.")
        
        # Check for missing values in critical columns
        if df[['Date and Time', 'Symbol', 'W Close']].isnull().any().any():
            raise ValueError("Missing data detected in critical columns.")
        
        # Convert 'Date and Time' to datetime, ensuring it is treated as date
        df['Date and Time'] = pd.to_datetime(df['Date and Time']).dt.date
        logging.info("Date and Time converted to datetime.")
        
        # Detect and delete duplicates based on 'Date and Time' and 'Symbol'
        duplicates = df.duplicated(subset=['Date and Time', 'Symbol'], keep='first')
        if duplicates.any():
            df.drop_duplicates(subset=['Date and Time', 'Symbol'], keep='first', inplace=True)
            logging.info(f"Removed {duplicates.sum()} duplicate entries.")
        
        # Sort by 'Date and Time' and 'Symbol'
        df.sort_values(by=['Symbol', 'Date and Time'], inplace=True)
        logging.info("Data sorted by 'Symbol' and 'Date and Time'.")
        
        # Filter out stocks with volume less than 0
        df = df[df['Volume'] >= 0]
        logging.info("Filtered out stocks with volume less than 50,000.")
        
        # Flag the first occurrence of each symbol
        df['Is New'] = ~df.duplicated(subset='Symbol', keep='first')
        logging.info("Flagged first occurrence of each symbol.")
        
        # Calculate daily returns for each symbol
        df['Daily Returns'] = df.groupby('Symbol')['W Close'].pct_change()
        logging.info("Calculated daily returns for each symbol.")
        
        # Calculate cumulative returns for each symbol using transform
        df['Cumulative Returns'] = df.groupby('Symbol')['Daily Returns'].transform(lambda x: (1 + x).cumprod())
        logging.info("Calculated cumulative returns for each symbol.")
        
        # Calculate holding days for each symbol
        df['First Date'] = df.groupby('Symbol')['Date and Time'].transform('min')
        df['First Date'] = pd.to_datetime(df['First Date']).dt.date
        df['Holding Days'] = (pd.to_datetime(df['Date and Time']) - pd.to_datetime(df['First Date'])).dt.days
        logging.info("Calculated holding days for each symbol.")
        
        return df
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        raise

def save_with_formatting(df, filename):
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Recommended_Symbols_Performance')
            workbook = writer.book
            
            # Get the worksheet
            worksheet = writer.sheets['Recommended_Symbols_Performance']
            
            # Define a green fill for new symbols
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            # Define a light yellow fill for today's analysis stocks
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            # Define a red fill for sell or strong sell symbols
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            # Get the last analysis date
            last_analysis_date = df['Date and Time'].max()
            
            # Apply green fill to rows with new symbols, yellow fill to today's analysis stocks,
            # and red fill to sell or strong sell symbols with the last or more recent analysis date
            for idx, (is_new, analysis_date, summary) in enumerate(zip(df['Is New'], df['Date and Time'], df['Summary']), start=1):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=idx + 1, column=col)
                    if is_new:
                        cell.fill = green_fill
                    elif analysis_date == date.today():
                        cell.fill = yellow_fill
                    elif analysis_date == last_analysis_date and summary.upper() in ['SELL', 'STRONG_SELL','NEUTRAL']:
                        cell.fill = red_fill
            
            logging.info("Applied formatting to new, today's analysis, and sell/strong sell symbols with the last or more recent analysis date.")
            
            # Create a new DataFrame for the analysis sheet
            analysis_df = df.groupby('Symbol').agg(
                First_Date=('Date and Time', 'min'),
                Summary_First=('Summary', 'first'),
                Last_Date=('Date and Time', 'max'),
                First_Day_Price=('W Close', 'first'),
                Last_Day_Price=('W Close', 'last'),
                Summary_Last=('Summary', 'last'),
                Holing_Days=('Holding Days', 'last'),
                Volume=('Volume', 'last'),
                
                Cumulative_Return=('Cumulative Returns', 'last')
            ).reset_index()
            
            # Write the analysis DataFrame to a new sheet
            analysis_df.to_excel(writer, index=False, sheet_name='Symbols_Analysis')
            logging.info("Analysis sheet created successfully.")
            
            # Apply conditional formatting for Cumulative Return
            analysis_worksheet = writer.sheets['Symbols_Analysis']
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            for idx, cumulative_return in enumerate(analysis_df['Cumulative_Return'], start=2):  # start=2 to skip the header
                cell = analysis_worksheet.cell(row=idx, column=10)  # Cumulative Return is in the 6th column
                if cumulative_return >= 1:
                    cell.fill = green_fill
                elif cumulative_return <= 1:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill            
            
            logging.info("Applied conditional formatting for Cumulative Return.")
    except Exception as e:
        logging.error(f"An error occurred while saving the file: {e}")
        raise

# File paths
input_file_path = 'W-Advance_Technical_Analysis_KMI100_STRONG_BUY.xlsx'
output_file_path = 'W-KMI100_Performance_Analysis.xlsx'

try:
    stock_data = read_and_analyze_stock_data(input_file_path)
    save_with_formatting(stock_data, output_file_path)
    print("Performance calculations completed successfully.")
except Exception as e:
    logging.error(f"Processing failed: {e}")
