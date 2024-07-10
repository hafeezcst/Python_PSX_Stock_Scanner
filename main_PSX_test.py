import datetime
import logging
import os
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import sqlalchemy
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

from analysis_functions import analyze_symbol
from convert_excel_to_lists import CUSTUM, KMIALL, KMI100, KMI30, MYLIST, QSE, CRYPTO
from email_functions import send_email
from telegram_message import send_telegram_message

# Set up logging
logging.basicConfig(filename='analysis_log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Global constants
VOLUME_THRESHOLD = 100000
MIN_VOLUME = 5000
AO_THRESHOLD = 0
RSI_THRESHOLD = 30

def get_analysis_type():
    return "D"  # Placeholder function

def get_recommendation_filter():
    return "STRONG_BUY"  # Placeholder function

def get_symbol_selection(country_selection):
    if country_selection == "P":
        return "KMI100"  # Placeholder symbol selection for Pakistan
    else:
        return "QSE"  # Placeholder symbol selection for Qatar

def main(country_selection):
    analysis_type = get_analysis_type()
    recommendation_filter = get_recommendation_filter()
    current_datetime = datetime.datetime.now()
    count = 1

    if country_selection == "P":
        screener_selection, exchange_selection = "PAKISTAN", "PSX"
        base_url = {
            'charts': "https://www.tradingview.com/chart/ZMYE714n/?symbol=PSX%3A",
            'finance': "https://www.tradingview.com/symbols/PSX-",
            'tech': "https://www.tradingview.com/symbols/PSX-"
        }
    else:
        screener_selection, exchange_selection = "QATAR", "QSE"
        base_url = {
            'charts': "https://www.tradingview.com/chart/ZMYE714n/?symbol=QSE%3A",
            'finance': "https://www.tradingview.com/symbols/QSE-",
            'tech': "https://www.tradingview.com/symbols/QSE-"
        }

    symbol_selection = get_symbol_selection(country_selection)
    symbols = {
        "KMI30": KMI30,
        "KMI100": KMI100,
        "MYLIST": MYLIST,
        "CRYPTO": CRYPTO,
        "CUSTUM": CUSTUM,
        "QSE": QSE
    }.get(symbol_selection, KMIALL)

    analyzed_data = []
    strong_buy_symbols = []
    buy_symbols = []
    sell_symbols = []
    AO_symbols = []

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [
            executor.submit(
                analyze_symbol, symbol, analysis_type, screener_selection, exchange_selection,
                base_url['charts'], base_url['finance'], base_url['tech']
            )
            for symbol in symbols
        ]

        for future in futures:
            result = future.result()
            if result:
                (symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume, ADX, RSI, RSI_Last, AO, 
                 Change, avg_support, avg_resistance, chart_url, finance_url, tech_url) = result
                
                summary = summary.get('RECOMMENDATION')
                print(f"{count}: Symbol: {symbol} - {summary}")
                print(f"{analysis_type} Sell_Signal:{Sell_Signal}, Neutral_Signal:{Neutral_Signal}, "
                      f"Buy_Signal:{Buy_Signal}, CLOSE: {Close}, Volume: {Volume}, ADX: {ADX}, RSI: {RSI}, "
                      f"LAST RSI: {RSI_Last}, AO: {AO}, %Change(D): {Change}\n")
                
                count += 1

                data_entry = [
                    current_datetime, symbol, summary, Close, Sell_Signal, Neutral_Signal, Buy_Signal, Volume,
                    ADX, RSI, RSI_Last, AO, Change, avg_support, avg_resistance, chart_url, finance_url, tech_url
                ]

                analyzed_data.append(data_entry)

                if (summary in ["BUY","NEUTRAL"] and isinstance(Volume, (int, float)) and Volume > VOLUME_THRESHOLD and isinstance(AO, (int, float)) and AO > AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI > 50):
                    strong_buy_symbols.append(data_entry)

                if summary == "NEUTRAL" and isinstance(Volume, (int, float)) and Volume > MIN_VOLUME and isinstance(AO, (int, float)) and AO > AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI > RSI_THRESHOLD:
                    buy_symbols.append(data_entry)

                if summary == "SELL" and isinstance(AO, (int, float)) and AO < AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI < 70:
                    sell_symbols.append(data_entry)

                if summary in ["NEUTRAL", "SELL", "STRONG_SELL"] and isinstance(AO, (int, float)) and AO > AO_THRESHOLD and isinstance(RSI, (int, float)) and RSI > RSI_THRESHOLD:
                    AO_symbols.append(data_entry)

    save_to_excel(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection, recommendation_filter)
    save_to_database(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection)

    send_notifications(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, analysis_type, symbol_selection, recommendation_filter)

def save_to_excel(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection, recommendation_filter):
    excel_file_path = f"{analysis_type}-Advance_Technical_Analysis_{symbol_selection}_{recommendation_filter}.xlsx"
    
    subset_columns = [
        'Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 
        'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance'
    ]
    subset_columns_drop = subset_columns[1:]

    dataframes = {
        'All Symbols': pd.DataFrame(analyzed_data, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'Recommended_Symbols': pd.DataFrame(strong_buy_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'Buy_Symbols': pd.DataFrame(buy_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'Sell_Symbols': pd.DataFrame(sell_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals']),
        'AO_Symbols': pd.DataFrame(AO_symbols, columns=subset_columns + ['Charts', 'Financials', 'Technicals'])
    }

    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        for sheet_name in dataframes:
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 20
            column_J_ranges = worksheet['J']
            column_L_ranges = worksheet['L']

            for cell in column_J_ranges:
                cell.font = Font(color="FF0000")

            for cell in column_L_ranges:
                cell.font = Font(color="FF0000")

            highlight_format_bullish = PatternFill(start_color="75AA74", end_color="75AA74", fill_type="solid")
            highlight_format_bearish = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            highlight_format_neutral = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            highlight_format_exit = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            dxf_bullish = DifferentialStyle(fill=highlight_format_bullish)
            dxf_bearish = DifferentialStyle(fill=highlight_format_bearish)
            dxf_neutral = DifferentialStyle(fill=highlight_format_neutral)
            dxf_exit = DifferentialStyle(fill=highlight_format_exit)

            rule_bullish = Rule(type="expression", dxf=dxf_bullish)
            rule_bullish.formula = ["$J1>50", "$L1>0"]
            rule_bearish = Rule(type="expression", dxf=dxf_bearish)
            rule_bearish.formula = ["$J1<50", "$L1<0"]
            rule_neutral = Rule(type="expression", dxf=dxf_neutral)
            rule_neutral.formula = ["$J1>50", "$L1>0"]
            rule_exit = Rule(type="expression", dxf=dxf_exit)
            rule_exit.formula = ["$E1>7"]

            worksheet.conditional_formatting.add("A1:R1000", rule_bullish)
            worksheet.conditional_formatting.add("E1:E1000", rule_exit)

def save_to_database(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, AO_symbols, analysis_type, symbol_selection):
    database_url = f'sqlite:///{symbol_selection}_{analysis_type}_data.db'
    engine = sqlalchemy.create_engine(database_url)

    dataframes = {
        'all_symbols': pd.DataFrame(analyzed_data, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'recommended_symbols': pd.DataFrame(strong_buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'buy_symbols': pd.DataFrame(buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'sell_symbols': pd.DataFrame(sell_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals']),
        'AO_symbols': pd.DataFrame(AO_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    }

    for table_name, df in dataframes.items():
        df.to_sql(table_name, engine, if_exists='append', index=False)

def send_notifications(analyzed_data, strong_buy_symbols, buy_symbols, sell_symbols, analysis_type, symbol_selection, recommendation_filter):
    today_date = datetime.datetime.now().date()
    
    df_strong_buy = pd.DataFrame(strong_buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    df_buy = pd.DataFrame(buy_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    df_sell = pd.DataFrame(sell_symbols, columns=['Date and Time', 'Symbol', 'Summary', f'{analysis_type} Close', 'Sell', 'Neutral', 'Buy', 'Volume', 'ADX', 'RSI', 'Last RSI', 'AO', '%Change(D)', 'Support', 'Resistance', 'Charts', 'Financials', 'Technicals'])
    
    today_Strong_buy = df_strong_buy[df_strong_buy['Date and Time'].dt.date == today_date]
    today_buy = df_buy[df_buy['Date and Time'].dt.date == today_date]
    today_sell = df_sell[df_sell['Date and Time'].dt.date == today_date]
    
    custom_message_strong_buy = f"Strong Buy Conditions: Volume >{VOLUME_THRESHOLD},RSI>{RSI_THRESHOLD}, AO >{AO_THRESHOLD} and {recommendation_filter} from Trading view."
    custom_message_buy = f"Buy Conditions: AO > {AO_THRESHOLD} and Buy from Trading view and RSI>{RSI_THRESHOLD}"
    custom_message_sell = f"Sell Conditions: AO < {AO_THRESHOLD} and Sell from Trading view and RSI<{RSI_THRESHOLD}"
    
    today_Strong_buy_mail = today_Strong_buy.to_string(index=True, justify='left', col_space=10)
    today_buy_mail = today_buy.to_string(index=True, justify='left', col_space=10)
    today_sell_mail = today_sell.to_string(index=True, justify='left', col_space=10)
    
    body = f"\n\n\n{custom_message_sell}\n{today_sell_mail}\n\n\n{custom_message_buy}\n{today_buy_mail}\n\n\n{custom_message_strong_buy}\n{today_Strong_buy_mail}"
    
    attachment_path = f"{analysis_type}-Advance_Technical_Analysis_{symbol_selection}_{recommendation_filter}.xlsx"
    subject = f"{analysis_type}-{symbol_selection}- {recommendation_filter}-Technical_Analysis_"
    
    try:
        send_email(subject, body, attachment_path)
        send_telegram_message(body)
    except Exception as e:
        print(f"Error sending email: {str(e)}")

def read_and_analyze_stock_data(file_path):
    try:
        # Read the Excel file from the specified sheet
        df = pd.read_excel(file_path, sheet_name='All Symbols')
        logging.info("Excel file read successfully.")
        
        # Check for missing values in critical columns
        if df[['Date and Time', 'Symbol', 'W Close']].isnull().any().any():
            raise ValueError("Missing data detected in critical columns.")
        
        # Convert 'Date and Time' to datetime, ensuring it is treated as date
        df['Date and Time'] = pd.to_datetime(df['Date and Time']).dt.date
        logging.info("Date and Time converted to datetime.")
        
        # Detect and delete duplicates based on 'Date and Time' and 'Symbol'
        duplicates = df.duplicated(subset=['Date and Time', 'Symbol'], keep='first')
        if duplicates.any():
            df.drop_duplicates(subset=['Date and Time', 'Symbol'], keep='first', inplace=True)
            logging.info(f"Removed {duplicates.sum()} duplicate entries.")
        
        # Sort by 'Date and Time' and 'Symbol'
        df.sort_values(by=['Symbol', 'Date and Time'], inplace=True)
        logging.info("Data sorted by 'Symbol' and 'Date and Time'.")
        
        # Filter out stocks with volume less than 0
        df = df[df['Volume'] >= 0]
        logging.info("Filtered out stocks with volume less than 0.")
        
        # Flag the first occurrence of each symbol
        df['Is New'] = ~df.duplicated(subset='Symbol', keep='first')
        logging.info("Flagged first occurrence of each symbol.")
        
        # Calculate daily returns for each symbol
        df['Daily Returns'] = df.groupby('Symbol')['W Close'].pct_change()
        logging.info("Calculated daily returns for each symbol.")
        
        # Calculate cumulative returns for each symbol using transform
        df['Cumulative Returns'] = df.groupby('Symbol')['Daily Returns'].transform(lambda x: (1 + x).cumprod())
        logging.info("Calculated cumulative returns for each symbol.")
        
        # Calculate holding days for each symbol
        df['First Date'] = df.groupby('Symbol')['Date and Time'].transform('min')
        df['First Date'] = pd.to_datetime(df['First Date']).dt.date
        df['Holding Days'] = (pd.to_datetime(df['Date and Time']) - pd.to_datetime(df['First Date'])).dt.days
        logging.info("Calculated holding days for each symbol.")
        
        return df
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        raise

def save_with_formatting(df, filename):
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Recommended_Symbols_Performance')
            workbook = writer.book
            
            # Get the worksheet
            worksheet = writer.sheets['Recommended_Symbols_Performance']
            
            # Define a green fill for new symbols
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            # Define a light yellow fill for today's analysis stocks
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            # Define a red fill for sell or strong sell symbols
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            # Get the last analysis date
            last_analysis_date = df['Date and Time'].max()
            
            # Apply green fill to rows with new symbols, yellow fill to today's analysis stocks,
            # and red fill to sell or strong sell symbols with the last or more recent analysis date
            for idx, (is_new, analysis_date, summary) in enumerate(zip(df['Is New'], df['Date and Time'], df['Summary']), start=1):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=idx + 1, column=col)
                    if is_new:
                        cell.fill = green_fill
                    elif analysis_date == date.today():
                        cell.fill = yellow_fill
                    elif analysis_date == last_analysis_date and summary.upper() in ['SELL', 'STRONG_SELL', 'NEUTRAL']:
                        cell.fill = red_fill
            
            logging.info("Applied formatting to new, today's analysis, and sell/strong sell symbols with the last or more recent analysis date.")
            
            # Create a new DataFrame for the analysis sheet
            analysis_df = df.groupby('Symbol').agg(
                First_Date=('Date and Time', 'min'),
                Summary_First=('Summary', 'first'),
                Last_Date=('Date and Time', 'max'),
                First_Day_Price=('W Close', 'first'),
                Last_Day_Price=('W Close', 'last'),
                Summary_Last=('Summary', 'last'),
                Holding_Days=('Holding Days', 'last'),
                Volume=('Volume', 'last'),
                Cumulative_Return=('Cumulative Returns', 'last')
            ).reset_index()
            
            # Write the analysis DataFrame to a new sheet
            analysis_df.to_excel(writer, index=False, sheet_name='Symbols_Analysis')
            logging.info("Analysis sheet created successfully.")
            
            # Apply conditional formatting for Cumulative Return
            analysis_worksheet = writer.sheets['Symbols_Analysis']
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            for idx, cumulative_return in enumerate(analysis_df['Cumulative_Return'], start=2):  # start=2 to skip the header
                cell = analysis_worksheet.cell(row=idx, column=10)  # Cumulative Return is in the 6th column
                if cumulative_return >= 1:
                    cell.fill = green_fill
                elif cumulative_return <= 1:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill            
            
            logging.info("Applied conditional formatting for Cumulative Return.")
    except Exception as e:
        logging.error(f"An error occurred while saving the file: {e}")
        raise

if __name__ == "__main__":
    current_day = datetime.datetime.now().weekday()
    current_month_day = datetime.datetime.now().day

    if current_day == 5:
        analysis_type = "W"
    elif current_month_day == 1:
        analysis_type = "M"
    else:
        analysis_type = "D"

    main("P")
    print("PSX Analysis complete")
    time.sleep(5)

    main("Q")
    print("QSE Analysis complete")
    time.sleep(5)

    print("All Analysis complete")
    time.sleep(5)

    # After initial analysis, read and further analyze the generated Excel file
    input_file_path = f"{analysis_type}-Advance_Technical_Analysis_KMI100_STRONG_BUY.xlsx"
    output_file_path = f"{analysis_type}-KMI100_Performance_Analysis.xlsx"

    try:
        stock_data = read_and_analyze_stock_data(input_file_path)
        save_with_formatting(stock_data, output_file_path)
        print("Performance calculations completed successfully.")
    except Exception as e:
        logging.error(f"Processing failed: {e}")
